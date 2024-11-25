from collections.abc import Sequence
import openpyxl
from openpyxl.utils import get_column_letter
from django.db import models
from django.db.models import Count, Sum, Q
from django.views import generic as g
from django.http.response import HttpResponseRedirect, HttpResponse
from django.urls import reverse_lazy, reverse
from django.contrib.auth.mixins import LoginRequiredMixin
from django.conf import settings
from django.http import JsonResponse
from django.shortcuts import redirect, get_object_or_404
from djmoney.money import Money
from django.db.models import F
from django.utils import timezone
from datetime import timedelta, datetime
from django.db.models.functions import ExtractDay, ExtractMonth, ExtractYear
from django.core.mail import send_mail
from django.contrib import messages
from django.views.decorators.csrf import csrf_exempt
from django.utils.decorators import method_decorator
from django.db.models.functions import TruncDay, TruncMonth

from apps.authentication.models import Department, User, User as Staff, UserRecoveryCode, reset_or_generate_code 
from apps.authentication.forms import AddDepartmentForm

from .models import Risk
from .forms import AddRiskForm, RiskFilterForm, AddRiskMinimalForm, AddStaffForm, UpdateStaffMinimalForm, UpdateStaffProfilePicForm, UpdateStaffImageForm,  StaffPasswordChangeForm
from .mixins import SuperUserMixin
from .utils import render_template, make_random_password


DEFAULT_PAGINATION_COUNT = settings.DEFAULT_PAGINATION_COUNT
DEFAULT_CURRENCY_SYMBOL = '₦'
MAX_SEVERE_RISKS = 10

from django.core.exceptions import PermissionDenied


from django.core.paginator import Paginator
from django.http import JsonResponse
from django.views.generic import ListView
from .models import Risk
from .forms import RiskFilterForm




def get_most_severe_risks(num, request=None):
    return Risk.get_risks(request).annotate(
            severity=F('probability') * F('impact')
        ).order_by('-severity')[:num]


def _make_aware(date):
    try:
        data = timezone.make_aware(
            date,
            timezone.get_current_timezone()
        )
    except Exception:
        data = date
    return data

def _from_iso(isodate):
    isodate = isodate[:-1] if isodate.upper().endswith('Z') else isodate
    return timezone.datetime.fromisoformat(isodate)



class Dashboard(SuperUserMixin, LoginRequiredMixin, g.TemplateView):
    template_name = "risk_manager/home.html"

    def get_context_data(self, **kwargs):
        """Insert the form into the context dict."""
        kwargs['severe_risks'] = get_most_severe_risks(MAX_SEVERE_RISKS, self.request)
        return super().get_context_data(**kwargs)


class MePasswordChange(SuperUserMixin, LoginRequiredMixin, g.TemplateView):
    template_name = "risk_manager/me/password_change.html"
    form_class = StaffPasswordChangeForm

    def get_context_data(self, **kwargs):
        """Insert the form into the context dict."""
        kwargs['form'] = self.get_form()
        kwargs['recovery_codes'] = self.request.user.recovery_codes.all()
        return super().get_context_data(**kwargs)

    def get_form(self):
        if self.request.method == "POST":
            return self.form_class(self.request.POST)
        return self.form_class()

    def post(self, request, *args, **kwargs):
        form = self.get_form()
        if form.is_valid():
            current_password = form.cleaned_data["current_password"]
            new_password = form.cleaned_data["new_password"]

            if not request.user.check_password(current_password):
                form.add_error("current_password", "The current password is incorrect.")
                return self.render_to_response(self.get_context_data(form=form))

            # Save the new password
            form.save(request.user)
            messages.success(request, "Your password has been successfully changed.")
            return redirect("risk_register:password_change")

        # If the form is invalid, re-render the page with errors
        return self.render_to_response(self.get_context_data(form=form))


class MeRecoveryCodeReset(SuperUserMixin, LoginRequiredMixin, g.View):
    def post(self, request, *args, **kwargs):
        me = self.request.user
        reset_or_generate_code(me)
        return redirect('risk_register:password_change')



class MeRecoveryCodeDownload(SuperUserMixin, LoginRequiredMixin, g.View):
    def get(self, request, *args, **kwargs):
        recovery_codes = UserRecoveryCode.objects.filter(user=request.user)
        if not recovery_codes.exists():
            return HttpResponse("No recovery codes found for download.", status=404)

        recovery_codes_text = "\n".join([code.code for code in recovery_codes])
        response = HttpResponse(recovery_codes_text, content_type="text/plain")
        response['Content-Disposition'] = 'attachment; filename="recovery_codes.txt"'
        
        return response


class DownloadRiskExcel(SuperUserMixin, LoginRequiredMixin, g.View):
    # TODO:
    #  use the filter form to generate the excel file

    def get_ordering(self):
        return ('-id', )
        
    def get(self, request):
        # Create an in-memory workbook and sheet
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Risks"

        # Set the headers
        headers = ["Risk Description", "Risk Type", "Risk Response", "Budget", "Probability", "Impact", "Rating", "Risk Owner", "Status", "Date Raised"]
        for col_num, header in enumerate(headers, 1):
            col_letter = get_column_letter(col_num)
            sheet[f"{col_letter}1"] = header

        # Populate the rows with risk data
        risks = RiskListView.do_risk_filter(self)  # Fetch your risk data (apply any filtering if necessary)
        for row_num, risk in enumerate(risks, 2):  # Start from row 2
            sheet[f"A{row_num}"] = risk.risk_description
            sheet[f"B{row_num}"] = risk.risk_type
            sheet[f"C{row_num}"] = risk.risk_response
            sheet[f"D{row_num}"] = f"₦{risk.risk_budget.amount if risk.risk_budget else 0}"
            sheet[f"E{row_num}"] = risk.get_prob_label()
            sheet[f"F{row_num}"] = risk.get_impact_label()
            sheet[f"G{row_num}"] = f"{risk.rating_percent()}%"
            sheet[f"H{row_num}"] = risk.risk_owner.name
            sheet[f"I{row_num}"] = "Closed" if risk.is_closed else "Opened"
            sheet[f"J{row_num}"] = risk.date_opened.strftime('%Y-%m-%d')

        # Prepare response for download
        response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response["Content-Disposition"] = 'attachment; filename="risk_list.xlsx"'
        workbook.save(response)
        return response



class AddRisk(SuperUserMixin, LoginRequiredMixin, g.CreateView):
    template_name = "risk_manager/add_risk.html"
    form_class = AddRiskForm
    model = Risk
    success_url = reverse_lazy('risk_register:risk_list')

    def form_valid(self, form):
        reponse = super().form_valid(form)
        self.object = form.save()
        response = HttpResponseRedirect(self.get_success_url())

        self.mail_others()
        
        return response

    def mail_others(self):
        # mail others in the department that the risk was added to
        risk = self.object
        me = self.request.user
        others = risk.risk_owner.staffs.exclude(id=me.id).values_list('email', flat=True)
        context = {
            'staff': self.request.user,
            'request': self.request,
            'risk': risk,
        }

        new_risk_me = render_template('risk_manager/email/new_risk_to_one.html', context)
        new_risk_others = render_template('risk_manager/email/new_risk_to_others.html', context)

        send_mail(
            'You successfully added a new risk',
            new_risk_me,
            settings.DEFAULT_FROM_EMAIL,
            [me.email],
            fail_silently=False
        )

        send_mail(
            'A new risk was added to oyour department',
            new_risk_others,
            settings.DEFAULT_FROM_EMAIL,
            others,
            fail_silently=False,
        )

    def form_invalid(self, form):
        """If the form
         is invalid, render the invalid form."""
        return self.render_to_response(self.get_context_data(form=form))

    def get_context_data(self, **kwargs):
        """Insert the form into the context dict."""
        if "form" not in kwargs:
            kwargs["form"] = self.get_form()
        return super().get_context_data(**kwargs)

    def get_form(self, *args, **kwargs):
        if not self.request.user.is_super_admin:
            self.form_class = AddRiskMinimalForm
        return super().get_form(*args, **kwargs)


class RiskListView(SuperUserMixin, LoginRequiredMixin, ListView):
    model = Risk
    template_name = 'risk_manager/risk_list.html'
    template_name_part = 'risk_manager/risk_list_part.html'
    template_name_part2 = 'risk_manager/risk_list_pagination_part.html'
    context_object_name = 'risks'
    paginate_by = settings.DEFAULT_PAGINATION_COUNT

    def get_queryset(self):
        return self.do_risk_filter(self)

    @staticmethod
    def do_risk_filter(self):
        # I know why not a mixin class, too lazy :)
        queryset = Risk.get_risks(self.request)
        self.filter_form = form = RiskFilterForm(self.request.GET)

        if form.is_valid():
            risk_type = form.cleaned_data.get('risk_type')
            probability = form.cleaned_data.get('probability')
            impact = form.cleaned_data.get('impact')
            is_closed = form.cleaned_data.get('is_closed')
            search_string = form.cleaned_data.get('search_string')

            if risk_type:
                queryset = queryset.filter(risk_type=risk_type)
            if probability:
                queryset = queryset.filter(probability=probability)
            if impact:
                queryset = queryset.filter(impact=impact)
            if is_closed:
                queryset = queryset.filter(is_closed=is_closed == 'True')
            if search_string:
                queryset = queryset.filter(
                    models.Q(risk_description__icontains=search_string) | 
                    models.Q(risk_type__icontains=search_string) | 
                    models.Q(risk_owner__name__icontains=search_string) |
                    models.Q(risk_owner__code__icontains=search_string) |
                    models.Q(risk_owner__description__icontains=search_string)
                )


        return queryset.order_by(*self.get_ordering())

    def get_ordering(self):
        return ('-id', )

    def get_paginated_data(self, queryset):
        paginator = Paginator(queryset, self.paginate_by)
        page_number = self.request.GET.get('page', 1)
        page_obj = paginator.get_page(page_number)

        # Include complete pagination context for AJAX handling
        data = {
            'risks': list(page_obj.object_list.values()),  # Convert queryset to list of dicts
            'page': page_obj.number,
            'pages': paginator.num_pages,
            'has_next': page_obj.has_next(),
            'has_previous': page_obj.has_previous(),
            'page_obj': page_obj,
            'paginator': paginator,
            'is_paginated': paginator.num_pages > 1,
        }
        self.data = data
        return data

    def get(self, request, *args, **kwargs):
        if self.is_fetched_request(request):
            # if using fetched request just return the html in parts, I could return json but that will mean more rendering logic on the frontend.
            self.template_name = self.template_name_part
            risk_html = super().get(request, *args, **kwargs).render().content.decode()
            self.get_paginated_data(self.object_list)
            pagination_html = render_template(self.template_name_part2, self.data)
            return JsonResponse({
                'risk_html': risk_html,
                'pagination_html': pagination_html,
            }, safe=False)

        return super().get(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['form'] = RiskFilterForm(self.request.GET)
        return context


class RiskDeleteView(SuperUserMixin, LoginRequiredMixin, g.View):
    model = Risk
    success_url = reverse_lazy('risk_register:risk_list')  # Replace with your desired redirect URL
    super_admin_only = True
    
    def get(self, request, pk):
        # the issue here is if users accidently navigate using get request data might be loss, but I don't want or need
        j = self.model.objects.all().filter(id=pk).delete()
        return redirect(self.success_url)


class RiskDetailView(SuperUserMixin, LoginRequiredMixin, g.UpdateView):
    model = Risk
    template_name = "risk_manager/risk_update.html"
    form_class = AddRiskForm
    context_object_name = 'risk'

    def get_success_url(self):
        # Redirect to the detail view of the updated risk after successful form submission
        return reverse_lazy('risk_register:risk_update', kwargs={'pk': self.object.pk})

    def get_context_data(self, **kwargs):
        """Insert the form into the context dict."""
        context = super().get_context_data(**kwargs)        
        context['form'] = self.form_class(instance=self.object)
        return context


class RiskStat(SuperUserMixin, LoginRequiredMixin, g.TemplateView):
    template_name = "risk_manager/risk_statistics.html"


class RiskPinned(SuperUserMixin, LoginRequiredMixin, g.TemplateView):
    template_name = "risk_manager/risk_pinned.html"


class RiskByDeptSummary(SuperUserMixin, LoginRequiredMixin, g.View):
    def get(self, request, *args, **kwargs):
        # Query to get counts of open and closed risks by department
        risk_distribution = (
            Risk.objects.values('risk_owner__name')
                .annotate(
                    open_count=Count('id', filter=Q(is_closed=False)),
                    closed_count=Count('id', filter=Q(is_closed=True))
                )
        )

        # return HttpResponse(risk_distribution)
        
        res = {
            'labels': [],
            'open_series': [],
            'closed_series': [],
        }

        for risk in risk_distribution:
            risk_owner = risk['risk_owner__name']
            open_count = risk['open_count']
            closed_count = risk['closed_count']
            
            res['labels'].append(risk_owner)
            res['open_series'].append(open_count)
            res['closed_series'].append(closed_count)

        return JsonResponse(res, safe=False)


class RiskPieSummary(SuperUserMixin, LoginRequiredMixin, g.View):
    # RiskByCategory
    def get(self, request, *args, **kwargs):
        risk_distribution = Risk.get_risks(request).values('risk_type').annotate(count=Count('id'))
        MAX_LEN_RISK_TYPE = 7
        
        res = {
            'series': [],
            'labels': [],
            'tooltips': []  # Include tooltips for full names
        }

        for risk in risk_distribution:
            risk_type, risk_count = risk['risk_type'], risk['count']
            short_risk_type = (risk_type[:MAX_LEN_RISK_TYPE] + '...') if len(risk_type) > MAX_LEN_RISK_TYPE else risk_type  # Truncate after MAX_LEN_RISK_TYPE chars
            
            res['series'].append(risk_count)
            # res['labels'].append(f'{short_risk_type} ({risk_count})')
            res['labels'].append(f'{risk_type} ({risk_count})')
            res['tooltips'].append(risk_type)  # Full risk type name for tooltip

        return JsonResponse(res, safe=False)


class RiskSuperSummaryView(SuperUserMixin, LoginRequiredMixin, g.View):
    def get(self, request, *args, **kwargs):
        # Aggregate the data
        opened_risks = Risk.get_risks(request).filter(is_closed=False)
        opened_risks_count = opened_risks.count()
        closed_risks_count = Risk.objects.filter(is_closed=True).count()
        
        # Sum the budgets for opened risks
        total_budget = opened_risks.aggregate(total_budget=Sum('risk_budget'))['total_budget'] or 0
        
        # Get currency symbol from the MoneyField
        budget_for_opened_risks_str = f"{DEFAULT_CURRENCY_SYMBOL}{total_budget:,.2f}" if total_budget else "₦0.00"

        # Prepare the response data
        response_data = {
            "opened_risks": opened_risks_count,
            "closed_risks": closed_risks_count,
            "budget_for_opened_risks": budget_for_opened_risks_str,
        }

        return JsonResponse(response_data)


class RiskSeveritySummaryView(SuperUserMixin, LoginRequiredMixin, g.View):
    def get(self, request, *args, **kwargs):
        # Get the top 10 most severe risks based on the rating
        most_severe_risks = get_most_severe_risks(num=10, request=request)

        # Prepare the response data
        response_data = []
        for risk in most_severe_risks:
            response_data.append({
                "risk_id": risk.id,
                "risk_type": risk.risk_type,
                "risk_description": risk.risk_description,
                "probability": risk.probability,
                "impact": risk.impact,
                "rating": risk.rating(), 
                "budget": f"{DEFAULT_CURRENCY_SYMBOL}{risk.risk_budget.amount:,.2f}" if risk.risk_budget else "₦0.00",
                "date_opened": risk.date_opened.isoformat(),
            })

        return JsonResponse(response_data, safe=False)


class RiskRatingSummary(SuperUserMixin, LoginRequiredMixin, g.View):
    def get(self, request, *args, **kwargs):
        risk_distribution = Risk.objects.annotate(
            rating=models.F('probability') * models.F('impact')
        ).values('rating').annotate(count=Count('id')).order_by('rating')
        
        res = {   
            'series': [],
            'labels': [],
        }

        for risk in risk_distribution:
            rating, risk_count = risk['rating'], risk['count']
            rating_info = Risk.rating_info(rating)
            label = f"{rating_info['tag']} ({risk_count})"
            res['series'].append(risk_count)
            res['labels'].append(label)

        return JsonResponse(res, safe=False)


# class RiskProgressChartView(SuperUserMixin, LoginRequiredMixin, g.View):
#     previos version had some issues with one day delay
#     def get(self, request, *args, **kwargs):
#         # Retrieve query parameters
#         view = request.GET.get('view', 'daily')
#         start_date = request.GET.get('startDate')
#         end_date = request.GET.get('endDate')

#         # Parse dates
#         start_date = _from_iso(start_date) if start_date else timezone.now() - timedelta(days=14)
#         end_date = _from_iso(end_date) if end_date else timezone.now()

#         start_date, end_date = min(start_date, end_date), max(start_date, end_date)

#         # Set date extraction logic based on view (daily or monthly)
#         labels = []
#         opened_risks_data = []
#         closed_risks_data = []

#         if view == 'daily':
#             days_range = (end_date - start_date).days + 1
#             labels = [(start_date + timedelta(days=i)).strftime('%d %b') for i in range(days_range)]
#             opened_risks_data = [0] * days_range
#             closed_risks_data = [0] * days_range

#             # Get opened risks and closed risks grouped by day
#             opened_risks = Risk.objects.filter(is_closed=False, date_opened__range=(start_date, end_date))\
#                 .annotate(day=ExtractDay('date_opened')) \
#                 .values('day').annotate(count=Count('id')).order_by('day')
#             closed_risks = Risk.objects.filter(is_closed=True, date_opened__range=(start_date, end_date))\
#                 .annotate(day=ExtractDay('date_opened')) \
#                 .values('day').annotate(count=Count('id')).order_by('day')

#             # Populate opened_risks_data and closed_risks_data
#             for entry in opened_risks:
#                 day = entry['day']
#                 index = (_make_aware(timezone.datetime(start_date.year, start_date.month, day)) - _make_aware(start_date)).days
#                 if 0 <= index < days_range:
#                     opened_risks_data[index] = entry['count']

#             for entry in closed_risks:
#                 day = entry['day']
#                 index = (_make_aware(timezone.datetime(start_date.year, start_date.month, day)) - _make_aware(start_date)).days
#                 if 0 <= index < days_range:
#                     closed_risks_data[index] = entry['count']

#         elif view == 'monthly':
#             months_range = ((end_date.year - start_date.year) * 12 + end_date.month - start_date.month + 1)
#             labels = [(start_date + timedelta(days=i*30)).strftime('%b %Y') for i in range(months_range)]
#             opened_risks_data = [0] * months_range
#             closed_risks_data = [0] * months_range

#             # Get opened risks and closed risks grouped by month and year
#             opened_risks = Risk.objects.filter(is_closed=False, date_opened__range=(start_date, end_date))\
#                 .annotate(month=ExtractMonth('date_opened'), year=ExtractYear('date_opened'))\
#                 .values('year', 'month').annotate(count=Count('id')).order_by('year', 'month')
#             closed_risks = Risk.objects.filter(is_closed=True, date_opened__range=(start_date, end_date))\
#                 .annotate(month=ExtractMonth('date_opened'), year=ExtractYear('date_opened'))\
#                 .values('year', 'month').annotate(count=Count('id')).order_by('year', 'month')

#             # Populate opened_risks_data and closed_risks_data
#             for entry in opened_risks:
#                 year = entry['year']
#                 month = entry['month']
#                 index = (year - start_date.year) * 12 + (month - start_date.month)
#                 if 0 <= index < months_range:
#                     opened_risks_data[index] = entry['count']

#             for entry in closed_risks:
#                 year = entry['year']
#                 month = entry['month']
#                 index = (year - start_date.year) * 12 + (month - start_date.month)
#                 if 0 <= index < months_range:
#                     closed_risks_data[index] = entry['count']

#         # Prepare the response data
#         response_data = {
#             'labels': labels,
#             'opened_risks': opened_risks_data,
#             'closed_risks': closed_risks_data,
#         }

#         return JsonResponse(response_data)




class RiskProgressChartView(SuperUserMixin, LoginRequiredMixin, g.View):
    def get(self, request, *args, **kwargs):
        # Retrieve query parameters
        view = request.GET.get('view', 'daily')
        start_date = request.GET.get('startDate')
        end_date = request.GET.get('endDate')

        # Parse and ensure timezone-aware dates
        start_date = _from_iso(start_date) if start_date else timezone.now() - timedelta(days=14)
        end_date = _from_iso(end_date) if end_date else timezone.now()
        start_date = timezone.make_aware(start_date) if timezone.is_naive(start_date) else start_date
        end_date = timezone.make_aware(end_date) if timezone.is_naive(end_date) else end_date

        start_date, end_date = min(start_date, end_date), max(start_date, end_date)

        # Initialize data arrays
        labels = []
        opened_risks_data = []
        closed_risks_data = []

        if view == 'daily':
            days_range = (end_date - start_date).days + 1
            labels = [(start_date + timedelta(days=i)).strftime('%d %b') for i in range(days_range)]
            opened_risks_data = [0] * days_range
            closed_risks_data = [0] * days_range

            # Query risks grouped by day
            opened_risks = (
                Risk.get_risks(request).filter(is_closed=False, date_opened__range=(start_date, end_date))
                .annotate(day=TruncDay('date_opened'))
                .values('day')
                .annotate(count=Count('id'))
                .order_by('day')
            )
            closed_risks = (
                Risk.get_risks(request).filter(is_closed=True, date_opened__range=(start_date, end_date))
                .annotate(day=TruncDay('date_opened'))
                .values('day')
                .annotate(count=Count('id'))
                .order_by('day')
            )

            # Populate data arrays
            for entry in opened_risks:
                day = entry['day'].date()
                index = (day - start_date.date()).days
                if 0 <= index < days_range:
                    opened_risks_data[index] = entry['count']

            for entry in closed_risks:
                day = entry['day'].date()
                index = (day - start_date.date()).days
                if 0 <= index < days_range:
                    closed_risks_data[index] = entry['count']

        elif view == 'monthly':
            months_range = ((end_date.year - start_date.year) * 12 + end_date.month - start_date.month + 1)
            labels = [(start_date + timedelta(days=i * 30)).strftime('%b %Y') for i in range(months_range)]
            opened_risks_data = [0] * months_range
            closed_risks_data = [0] * months_range

            # Query risks grouped by month and year
            opened_risks = (
                Risk.get_risks(request).filter(is_closed=False, date_opened__range=(start_date, end_date))
                .annotate(month=TruncMonth('date_opened'))
                .values('month')
                .annotate(count=Count('id'))
                .order_by('month')
            )
            closed_risks = (
                Risk.get_risks(request).filter(is_closed=True, date_opened__range=(start_date, end_date))
                .annotate(month=TruncMonth('date_opened'))
                .values('month')
                .annotate(count=Count('id'))
                .order_by('month')
            )

            # Populate data arrays
            for entry in opened_risks:
                month = entry['month'].date()
                index = (month.year - start_date.year) * 12 + (month.month - start_date.month)
                if 0 <= index < months_range:
                    opened_risks_data[index] = entry['count']

            for entry in closed_risks:
                month = entry['month'].date()
                index = (month.year - start_date.year) * 12 + (month.month - start_date.month)
                if 0 <= index < months_range:
                    closed_risks_data[index] = entry['count']

        # Prepare the response data
        response_data = {
            'labels': labels,
            'opened_risks': opened_risks_data,
            'closed_risks': closed_risks_data,
        }

        return JsonResponse(response_data)



class DepartmentListView(SuperUserMixin, LoginRequiredMixin, g.ListView):
    model = Department
    template_name = 'risk_manager/departments/list.html'
    context_object_name = 'departments'
    paginate_by = settings.DEFAULT_PAGINATION_COUNT  # Show 10 risks per page


class DepartmentAddView(SuperUserMixin, LoginRequiredMixin, g.CreateView):
    # TODO:
    # - instead of redirecting to the department list view go to the detail page after creating a department
    template_name = "risk_manager/departments/add.html"
    form_class = AddDepartmentForm
    model = Department
    success_url = reverse_lazy('risk_register:dept_list')
    super_admin_only = True

    def form_valid(self, form):
        reponse = super().form_valid(form)
        self.object = form.save()
        self.form = form
        return HttpResponseRedirect(self.get_success_url())

    def get_success_url(self):
        return reverse('risk_register:dept_read', kwargs={'pk': self.object.id})

    def get_context_data(self, **kwargs):
        """Insert the form into the context dict."""
        if "form" not in kwargs:
            kwargs["form"] = self.get_form()
        return super().get_context_data(**kwargs)



class DepartmentUpdateView(SuperUserMixin, LoginRequiredMixin, g.UpdateView):
    model = Department
    form_class = AddDepartmentForm
    template_name = "risk_manager/departments/details.html"
    context_object_name = 'department'

    def get_success_url(self):
        # Redirect to the detail view of the updated risk after successful form submission
        return reverse_lazy('risk_register:dept_update', kwargs={'pk': self.object.pk})

    def get_context_data(self, **kwargs):
        """Insert the form into the context dict."""
        context = super().get_context_data(**kwargs)        
        context['form'] = self.form_class(instance=self.object)
        return context

    def get_context_data(self, **kwargs):
        """Insert the form into the context dict."""
        if "form" not in kwargs:
            kwargs["form"] = self.get_form()
        department = self.get_object()
        staffs = department.staffs.order_by('is_super_admin')
        kwargs['staffs'] = staffs
        return super().get_context_data(**kwargs)

    def form_valid(self, form):
        # Save the staff object
        redirect_response = super().form_valid(form)
        if self.request.user.is_super_admin:
            self.object = form.save()
        else:
            return redirect('risk_register:page_403')
        return redirect_response



class DepartmentDeleteView(SuperUserMixin, LoginRequiredMixin, g.View):
    model = Department
    success_url = reverse_lazy('risk_register:dept_list')  # Replace with your desired redirect URL
    super_admin_only = True
    
    def get(self, request, pk):
        # the issue here is if users accidently navigate using get request data might be loss, but I don't want or need
        j = self.model.objects.all().filter(id=pk).delete()
        return redirect(self.success_url)


class StaffAddView(SuperUserMixin, LoginRequiredMixin, g.CreateView):
    template_name = "risk_manager/staffs/staff_add.html"
    form_class = AddStaffForm
    model = Staff
    success_url = reverse_lazy('risk_register:staff_list')
    super_admin_only = True

    def form_valid(self, form):
        # Save the staff object
        response = super().form_valid(form)
        self.object = new_staff = form.save(commit=False)
        
        # Generate a random password
        # password = make_random_password(settings.MAX_RANDOM_PASSWORD_LENGTH, settings.RANDOM_PASSWORD_ALLOWED_CHARS)
        password = new_staff.last_name.upper()
        new_staff.set_password(password)
        new_staff.save()

        # context = {
        #     'password': password,
        #     'user': new_staff,
        #     'new_staff': new_staff,
        # }
        
        # your_password_html = render_template('risk_manager/email/account_created.html', context)
        # new_staff_html = render_template('risk_manager/email/new_staff.html', context)
        # others_in_my_department = new_staff.department.staffs.exclude(id=new_staff.id).values_list('email', flat=True)

        # send_mail(
        #     'Your Staff Account Password',
        #     your_password_html,
        #     settings.DEFAULT_FROM_EMAIL,
        #     [new_staff.email],
        #     fail_silently=False,
        # )

        # send_mail(
        #     'New Staff',
        #     new_staff_html,
        #     settings.DEFAULT_FROM_EMAIL,
        #     others_in_my_department,
        #     fail_silently=False,
        # )
        
        return response

    def form_invalid(self, form):
        """If the form
         is invalid, render the invalid form."""
        return self.render_to_response(self.get_context_data(form=form))

    def get_context_data(self, **kwargs):
        """Insert the form into the context dict."""
        if "form" not in kwargs:
            kwargs["form"] = self.get_form()
        return super().get_context_data(**kwargs)


class StaffDeleteView(SuperUserMixin, LoginRequiredMixin, g.View):
    model = Staff
    success_url = reverse_lazy('risk_register:staff_list')  # Replace with your desired redirect URL
    super_admin_only = True
    
    def get(self, request, pk):
        # the issue here is if users accidently navigate using get request data might be loss, but I don't want or need
        user = self.model.objects.all().filter(id=pk).first()
        if user == request.user:
            return HttpResponse('Error! You tried deleting yourself.')
        else:
            user.delete()
        return redirect(self.success_url)


class StaffListView(SuperUserMixin, LoginRequiredMixin, g.ListView):
    template_name = "risk_manager/staffs/staff_list.html"
    model = Staff
    context_object_name = 'staffs'
    paginate_by = settings.DEFAULT_PAGINATION_COUNT

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # context['form'] = RiskFilterForm(self.request.GET)
        return context

    def get_queryset(self):
        queryset = super().get_queryset()
        current_user = self.request.user

        if queryset.filter(id=current_user.id).exists():
            # Reorder the queryset with the current user first
            queryset = queryset.exclude(id=current_user.id)
            queryset = Staff.objects.filter(id=current_user.id).union(queryset)
        
        return queryset.order_by('-date_joined')


class StaffUpdateView(SuperUserMixin, LoginRequiredMixin, g.UpdateView):
    model = Staff
    template_name = "risk_manager/staffs/staff_update.html"
    form_class = AddStaffForm
    context_object_name = 'staff'

    def get_success_url(self):
        # Redirect to the detail view of the updated risk after successful form submission
        me = self.request.user
        staff = self.get_object()
        # if me == staff:
        #     # a staff editing his/herself should be redirected to either the update page or user page if any
        #     return reverse_lazy('risk_register:staff_update', kwargs={'pk': me.id})
        # return reverse_lazy('risk_register:staff_list')
        messages.success(self.request, "User updated successfully.")
        return reverse_lazy('risk_register:staff_update', kwargs={'pk': staff.id})

    def get_context_data(self, **kwargs):
        """Insert the form into the context dict."""
        context = super().get_context_data(**kwargs)        
        # context['form'] = self.form_class(instance=self.object)
        context['profile_pic_form'] = UpdateStaffProfilePicForm(instance=self.object)
        context['profile_image_form'] = UpdateStaffImageForm(instance=self.object)
        return context

    def dispatch(self, request, pk):
        # only super users can edit other staffs
        if not request.user.is_super_admin and pk != request.user.id:
            return redirect('risk_register:page_403')
        return super().dispatch(request, pk)

    def get_form(self, *args, **kwargs):
        if self.request.user.is_super_admin:
            self.form_class = AddStaffForm
        else:
            self.form_class = UpdateStaffMinimalForm
        return super().get_form(*args, **kwargs)



class UpdateStaffProfilePicView(SuperUserMixin, LoginRequiredMixin, g.View):
    model = Staff
    form_class = UpdateStaffProfilePicForm

    @method_decorator(csrf_exempt)
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)

    def post(self, request, *a, **kw):
        form = self.form_class(request.POST)

        if form.is_valid():
            profile_pic = form.cleaned_data.get('profile_pic')
            if profile_pic:
                me = request.user
                me.profile_pic = profile_pic
                me.uploaded_profile_pic = None
                me.save()

        return redirect('risk_register:staff_update', pk=me.id)



class UpdateUploadedProfilePicView(SuperUserMixin, LoginRequiredMixin, g.View):
    model = User
    form_class = UpdateStaffImageForm

    @method_decorator(csrf_exempt)
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)

    def post(self, request, *a, **kw):
        form = self.form_class(request.POST, request.FILES)  # Include request.FILES for file uploads

        if form.is_valid():
            uploaded_profile_pic = form.cleaned_data.get('uploaded_profile_pic')
            if uploaded_profile_pic:
                me = request.user
                me.uploaded_profile_pic = uploaded_profile_pic
                me.save()

        return redirect('risk_register:staff_update', pk=request.user.id)



class Page403(LoginRequiredMixin, g.TemplateView):
    template_name = 'risk_manager/403.html'


